from openpyxl import Workbook

from core.apply_excel import find_row, resolve_column_in_block, resolve_mvd_block, resolve_mvd_column


def test_resolve_mvd_column() -> None:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Р."
    ws["E1"] = "МВД"
    assert resolve_mvd_column(ws) == 5


def test_find_row_with_code() -> None:
    wb = Workbook()
    ws = wb.active
    ws["A10"] = "т. 2.5."
    ws["B10"] = "Задержание лиц, подозреваемых"

    row, suggestions, match_count = find_row(
        ws,
        row_label=None,
        row_label_contains="Задержание лиц",
        row_label_regex=None,
        row_label_column="B",
        row_code="т. 2.5.",
        row_code_column="A",
    )
    assert row == 10
    assert match_count == 1
    assert suggestions


def test_resolve_mvd_block() -> None:
    wb = Workbook()
    ws = wb.active
    ws.merge_cells("H1:K1")
    ws["H1"] = "МВД"
    ws["H2"] = "ЗАРЕГ. (дел) 2"
    ws["K2"] = "ПРЕКРАЩ. (дел) 2"

    block = resolve_mvd_block(ws)
    assert block == [8, 9, 10, 11]

    col = resolve_column_in_block(ws, block, "ЗАРЕГ.")
    assert col == 8
