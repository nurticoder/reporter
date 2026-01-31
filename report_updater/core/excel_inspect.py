from __future__ import annotations

from collections import Counter
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .apply_excel import resolve_mvd_block, resolve_mvd_column, resolve_sheet
from .normalize import normalize_label


def _header_snapshot(ws, max_rows: int = 3) -> list[dict]:
    rows: list[dict] = []
    for row in range(1, max_rows + 1):
        values = []
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=row, column=col).value
            if value is None or str(value).strip() == "":
                continue
            values.append({"col": get_column_letter(col), "value": value})
        if values:
            rows.append({"row": row, "values": values})
    return rows


def _sheet_label_stats(ws, label_col: str = "B") -> dict:
    labels: list[str] = []
    rows: list[tuple[int, str]] = []
    for row in range(1, ws.max_row + 1):
        value = ws[f"{label_col}{row}"].value
        if isinstance(value, str) and value.strip():
            labels.append(value.strip())
            rows.append((row, value.strip()))

    counts = Counter(labels)
    duplicates = [(label, count) for label, count in counts.items() if count > 1]
    duplicates.sort(key=lambda item: -item[1])

    return {
        "label_column": label_col,
        "label_count": len(labels),
        "duplicate_count": len(duplicates),
        "duplicate_samples": duplicates[:5],
        "sample_labels": rows[:10],
    }


def inspect_excel(excel_path: str) -> dict[str, Any]:
    workbook = load_workbook(excel_path, data_only=True)
    summaries: list[dict[str, Any]] = []
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        summaries.append(
            {
                "sheet": sheet,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "headers": _header_snapshot(ws),
            }
        )

    details: dict[str, Any] = {}
    sheet_r2 = resolve_sheet(workbook, "Р.2")
    if sheet_r2:
        ws = workbook[sheet_r2]
        details["r2"] = {
            "sheet": sheet_r2,
            "mvd_col": resolve_mvd_column(ws),
            "labels": _sheet_label_stats(ws, "B"),
            "row_code_column": "A",
        }

    sheet_r1 = resolve_sheet(workbook, "Отчет 1-Е Р.1") or resolve_sheet(workbook, "Р.1")
    if sheet_r1:
        ws = workbook[sheet_r1]
        block_cols = resolve_mvd_block(ws)
        subheaders = {}
        if block_cols:
            for col in block_cols:
                value = ws.cell(row=2, column=col).value
                if isinstance(value, str) and value.strip():
                    subheaders[get_column_letter(col)] = value.strip()
        details["r1"] = {
            "sheet": sheet_r1,
            "mvd_block": [get_column_letter(col) for col in block_cols] if block_cols else [],
            "subheaders": subheaders,
            "labels": _sheet_label_stats(ws, "B"),
        }

    return {"sheets": summaries, "details": details}
