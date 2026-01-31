from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .apply_excel import resolve_mvd_block, resolve_mvd_column, resolve_sheet
from .normalize import normalize_label


def diff_template(prev_xlsx: str, current_xlsx: str) -> list[dict]:
    prev = load_workbook(prev_xlsx, data_only=True)
    curr = load_workbook(current_xlsx, data_only=True)

    diffs: list[dict] = []
    for sheet in prev.sheetnames:
        if sheet not in curr.sheetnames:
            continue
        ws_prev = prev[sheet]
        ws_curr = curr[sheet]

        if normalize_label(sheet) == normalize_label("Р.2"):
            col = resolve_mvd_column(ws_prev)
            if not col:
                continue
            for row in range(1, ws_prev.max_row + 1):
                v_prev = ws_prev.cell(row=row, column=col).value
                v_curr = ws_curr.cell(row=row, column=col).value
                if v_prev != v_curr:
                    diffs.append(
                        {
                            "sheet": sheet,
                            "row": row,
                            "col": col,
                            "old": v_prev,
                            "new": v_curr,
                        }
                    )

        if normalize_label(sheet) == normalize_label("Отчет 1-Е Р.1"):
            block_prev = resolve_mvd_block(ws_prev)
            block_curr = resolve_mvd_block(ws_curr)
            if not block_prev or not block_curr:
                continue
            for row in range(1, ws_prev.max_row + 1):
                for col in block_prev:
                    v_prev = ws_prev.cell(row=row, column=col).value
                    v_curr = ws_curr.cell(row=row, column=col).value
                    if v_prev != v_curr:
                        diffs.append(
                            {
                                "sheet": sheet,
                                "row": row,
                                "col": col,
                                "old": v_prev,
                                "new": v_curr,
                            }
                        )

    return diffs
