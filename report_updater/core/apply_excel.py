from __future__ import annotations

import difflib
from dataclasses import dataclass
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .normalize import normalize_label, parse_int


@dataclass
class TargetResolution:
    sheet: str
    row: int
    col: int
    row_label: str
    col_header: str


def resolve_sheet(workbook, desired: str) -> str | None:
    if desired in workbook.sheetnames:
        return desired
    target = normalize_label(desired)
    candidates = []
    for name in workbook.sheetnames:
        normalized = normalize_label(name)
        if normalized == target:
            return name
        if target and (normalized.endswith(target) or target.endswith(normalized) or target in normalized):
            candidates.append(name)
    if len(candidates) == 1:
        return candidates[0]
    return None


def get_row_labels(ws, col_letter: str) -> list[tuple[int, str]]:
    labels = []
    for row in range(1, ws.max_row + 1):
        value = ws[f"{col_letter}{row}"].value
        if isinstance(value, str) and value.strip():
            labels.append((row, value.strip()))
    return labels


def suggest_rows(labels: list[tuple[int, str]], target: str) -> list[dict]:
    normalized_target = normalize_label(target)
    scored = []
    for row, label in labels:
        score = difflib.SequenceMatcher(None, normalize_label(label), normalized_target).ratio()
        scored.append((score, row, label))
    scored.sort(reverse=True)
    return [{"row": row, "label": label} for score, row, label in scored[:5]]


def find_row(
    ws,
    row_label: str | None,
    row_label_contains: str | None,
    row_label_regex: str | None,
    row_label_column: str,
    row_code: str | None,
    row_code_column: str,
) -> tuple[int | None, list[dict], int]:
    labels = get_row_labels(ws, row_label_column)
    matches = []

    for row, label in labels:
        if row_code:
            code_value = ws[f"{row_code_column}{row}"].value
            if not isinstance(code_value, str) or normalize_label(code_value) != normalize_label(row_code):
                continue
        if row_label_regex:
            try:
                if row_label_regex and __import__("re").search(row_label_regex, str(label), __import__("re").IGNORECASE):
                    matches.append((row, label))
            except Exception:
                continue
        elif row_label_contains:
            if normalize_label(row_label_contains) in normalize_label(label):
                matches.append((row, label))
        elif row_label:
            if normalize_label(label) == normalize_label(row_label):
                matches.append((row, label))

    suggestions = suggest_rows(labels, row_label or row_label_contains or "")
    if len(matches) == 1:
        return matches[0][0], suggestions, 1
    if len(matches) > 1:
        return None, suggestions, len(matches)
    return None, suggestions, 0


def find_header_row(ws, header_text: str, max_rows: int = 6) -> int | None:
    for row in range(1, max_rows + 1):
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=row, column=col).value
            if isinstance(value, str) and normalize_label(value) == normalize_label(header_text):
                return row
    return None


def resolve_mvd_column(ws) -> int | None:
    for row in range(1, 6):
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=row, column=col).value
            if isinstance(value, str) and normalize_label(value) == normalize_label("МВД"):
                return col
    return None


def resolve_mvd_block(ws) -> list[int]:
    mvd_cell = None
    for row in range(1, 6):
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=row, column=col).value
            if isinstance(value, str) and normalize_label(value) == normalize_label("МВД"):
                mvd_cell = (row, col)
                break
        if mvd_cell:
            break
    if not mvd_cell:
        return []

    # check merged ranges
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= mvd_cell[0] <= merged.max_row and merged.min_col <= mvd_cell[1] <= merged.max_col:
            return list(range(merged.min_col, merged.max_col + 1))

    # fallback to 4 columns block
    return list(range(mvd_cell[1], mvd_cell[1] + 4))


def resolve_column_in_block(ws, block_cols: list[int], header_contains: str) -> int | None:
    header_row = None
    for row in range(1, 6):
        row_values = [ws.cell(row=row, column=c).value for c in block_cols]
        if any(isinstance(v, str) and v.strip() for v in row_values):
            header_row = row
    if header_row is None:
        return None

    target = normalize_label(header_contains)
    for col in block_cols:
        value = ws.cell(row=header_row, column=col).value
        if isinstance(value, str) and target in normalize_label(value):
            return col
    return None


def plan_metric_updates(workbook, metrics: dict, excel_map: dict) -> tuple[list[dict], list[dict], list[dict]]:
    updates = []
    errors = []
    warnings = []
    debug = []

    for metric_key, rule in excel_map.items():
        if metric_key not in metrics:
            continue
        targets = rule.get("targets", [])
        for target in targets:
            sheet_name = target.get("sheet")
            resolved_sheet = resolve_sheet(workbook, sheet_name)
            if not resolved_sheet:
                errors.append(
                    {
                        "type": "error",
                        "message": f"Лист '{sheet_name}' не найден для метрики {metric_key}.",
                        "source": "excel_map.yaml",
                        "suggestedFix": "Проверьте имя листа в excel_map.yaml.",
                    }
                )
                continue

            ws = workbook[resolved_sheet]
            row, suggestions, match_count = find_row(
                ws,
                target.get("row_label"),
                target.get("row_label_contains"),
                target.get("row_label_regex"),
                target.get("row_label_column", "B"),
                target.get("row_code"),
                target.get("row_code_column", "A"),
            )
            if row is None:
                if match_count > 1:
                    errors.append(
                        {
                            "type": "error",
                            "message": f"Неоднозначное совпадение строк для метрики {metric_key}.",
                            "source": resolved_sheet,
                            "suggestedFix": "Уточните row_label_contains/regex или добавьте row_code.",
                            "candidates": suggestions,
                        }
                    )
                    continue
                errors.append(
                    {
                        "type": "error",
                        "message": f"Не найдена строка для метрики {metric_key}.",
                        "source": resolved_sheet,
                        "suggestedFix": "Уточните row_label_contains или row_label_regex.",
                        "candidates": suggestions,
                    }
                )
                continue

            col_key = target.get("col_key", "МВД")
            col = resolve_mvd_column(ws) if normalize_label(col_key) == normalize_label("МВД") else None
            if col is None:
                errors.append(
                    {
                        "type": "error",
                        "message": f"Колонка '{col_key}' не найдена для метрики {metric_key}.",
                        "source": resolved_sheet,
                        "suggestedFix": "Проверьте заголовок колонки в Excel.",
                    }
                )
                continue

            cell = ws.cell(row=row, column=col)
            updates.append(
                {
                    "sheet": resolved_sheet,
                    "cell": f"{get_column_letter(col)}{row}",
                    "rowLabel": ws[f"{target.get('row_label_column','B')}{row}"].value,
                    "oldValue": cell.value,
                    "newValue": metrics[metric_key]["value"],
                    "kind": metric_key,
                    "row": row,
                    "col": col,
                }
            )
            debug.append(
                {
                    "metric": metric_key,
                    "sheet": resolved_sheet,
                    "row": row,
                    "col": col,
                    "cell": f"{get_column_letter(col)}{row}",
                }
            )

    return updates, errors, warnings, debug


def plan_article_updates(workbook, article_breakdown: list[dict], article_map: dict) -> tuple[list[dict], list[dict], list[dict], list[dict]]:
    updates = []
    errors = []
    warnings = []
    debug = []

    sheet_name = article_map.get("sheet")
    if not sheet_name:
        return updates, errors, warnings, debug

    resolved_sheet = resolve_sheet(workbook, sheet_name)
    if not resolved_sheet:
        errors.append(
            {
                "type": "error",
                "message": f"Лист '{sheet_name}' не найден для статей.",
                "source": "article_map.yaml",
                "suggestedFix": "Проверьте имя листа в article_map.yaml.",
            }
        )
        return updates, errors, warnings, debug

    ws = workbook[resolved_sheet]
    block_cols = resolve_mvd_block(ws)
    if not block_cols:
        errors.append(
            {
                "type": "error",
                "message": "Блок МВД не найден на листе статей.",
                "source": resolved_sheet,
                "suggestedFix": "Проверьте заголовки в Excel.",
            }
        )
        return updates, errors, warnings, debug

    row_label_column = article_map.get("row_label_column", "B")
    fields = article_map.get("fields", {})

    for row in article_breakdown:
        article = row.get("article")
        if not article:
            continue
        excel_row, suggestions, match_count = find_row(
            ws,
            article,
            article_map.get("row_label_contains"),
            article_map.get("row_label_regex"),
            row_label_column,
            None,
            "A",
        )
        if excel_row is None:
            if match_count > 1:
                errors.append(
                    {
                        "type": "error",
                        "message": f"Неоднозначное совпадение для статьи '{article}'.",
                        "source": resolved_sheet,
                        "suggestedFix": "Уточните regex для строки статьи.",
                        "candidates": suggestions,
                    }
                )
                continue
            errors.append(
                {
                    "type": "error",
                    "message": f"Статья '{article}' не найдена в Excel.",
                    "source": resolved_sheet,
                    "suggestedFix": "Добавьте строку статьи или уточните regex.",
                    "candidates": suggestions,
                }
            )
            continue

        for field_key, field_rule in fields.items():
            if field_key not in row:
                continue
            col = resolve_column_in_block(ws, block_cols, field_rule.get("header_contains", ""))
            if col is None:
                errors.append(
                    {
                        "type": "error",
                        "message": f"Колонка для поля '{field_key}' не найдена в блоке МВД.",
                        "source": resolved_sheet,
                        "suggestedFix": "Проверьте заголовки под блоком МВД.",
                    }
                )
                continue
            cell = ws.cell(row=excel_row, column=col)
            updates.append(
                {
                    "sheet": resolved_sheet,
                    "cell": f"{get_column_letter(col)}{excel_row}",
                    "rowLabel": article,
                    "oldValue": cell.value,
                    "newValue": row[field_key],
                    "kind": f"article_{field_key}",
                    "row": excel_row,
                    "col": col,
                }
            )
            debug.append(
                {
                    "article": article,
                    "field": field_key,
                    "sheet": resolved_sheet,
                    "row": excel_row,
                    "col": col,
                }
            )

    return updates, errors, warnings, debug


def plan_updates(excel_path: str, metrics: dict, article_breakdown: list[dict], excel_map: dict, article_map: dict):
    workbook = load_workbook(excel_path)
    metric_updates, metric_errors, metric_warnings, metric_debug = plan_metric_updates(
        workbook, metrics, excel_map
    )
    article_updates, article_errors, article_warnings, article_debug = plan_article_updates(
        workbook, article_breakdown, article_map
    )

    updates = metric_updates + article_updates
    errors = metric_errors + article_errors
    warnings = metric_warnings + article_warnings
    debug = metric_debug + article_debug

    # detect conflicts
    seen = {}
    for update in updates:
        key = (update["sheet"], update["row"], update["col"])
        if key in seen and seen[key] != update["newValue"]:
            errors.append(
                {
                    "type": "error",
                    "message": f"Конфликт записи в ячейку {update['cell']}.",
                    "source": update["sheet"],
                    "suggestedFix": "Проверьте маппинг метрик, две метрики пишут в одну ячейку.",
                }
            )
        seen[key] = update["newValue"]

    return updates, errors, warnings, debug


def apply_updates(
    excel_path: str,
    output_path: str,
    updates: list[dict],
) -> None:
    workbook = load_workbook(excel_path)
    for update in updates:
        ws = workbook[update["sheet"]]
        ws.cell(row=update["row"], column=update["col"], value=update["newValue"])
    workbook.save(output_path)
