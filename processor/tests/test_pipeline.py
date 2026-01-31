from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from config_loader import (
    load_article_map,
    load_cross_checks,
    load_excel_map,
    load_metric_dictionary,
    load_required_metrics,
)
from extract_word import extract_word_data
from update_excel import apply_updates, find_row_by_label, plan_updates
from validate import validate_report
from .fixtures_factory import create_sample_docx, create_sample_docx_missing_metric, create_sample_xlsx


def run_validation(word_path: Path, excel_path: Path):
    metric_dictionary = load_metric_dictionary()
    required_metrics = load_required_metrics()
    excel_map = load_excel_map()
    article_map = load_article_map()
    cross_checks = load_cross_checks()

    word_data = extract_word_data(str(word_path), metric_dictionary)
    validation = validate_report(
        report_month=word_data["report_month"],
        metrics=word_data["metrics"],
        metrics_list=word_data["metrics_list"],
        cases=word_data["cases"],
        required_metrics=required_metrics,
        cross_checks=cross_checks,
        article_breakdown=word_data["article_breakdown"],
        article_map=article_map,
        issues=word_data["metric_issues"] + word_data["case_issues"],
    )

    updates, plan_errors = plan_updates(
        str(excel_path),
        word_data["metrics"],
        word_data["article_breakdown"],
        excel_map,
        article_map,
    )
    if plan_errors:
        validation["errors"].extend(plan_errors)

    return word_data, validation, updates


def test_extract_validate_success(tmp_path: Path):
    word_path = tmp_path / "sample.docx"
    excel_path = tmp_path / "sample.xlsx"
    create_sample_docx(word_path)
    create_sample_xlsx(excel_path)

    _, validation, _ = run_validation(word_path, excel_path)

    assert validation["errors"] == []
    assert any(item["key"] == "remaining_cases" for item in validation["cross_checks"])


def test_missing_metric_blocks(tmp_path: Path):
    word_path = tmp_path / "missing.docx"
    excel_path = tmp_path / "sample.xlsx"
    create_sample_docx_missing_metric(word_path)
    create_sample_xlsx(excel_path)

    _, validation, _ = run_validation(word_path, excel_path)

    assert any("Missing required metric" in error["message"] for error in validation["errors"])


def test_apply_updates_changes_cells(tmp_path: Path):
    word_path = tmp_path / "sample.docx"
    excel_path = tmp_path / "sample.xlsx"
    output_path = tmp_path / "output.xlsx"
    create_sample_docx(word_path)
    create_sample_xlsx(excel_path)

    word_data, validation, updates = run_validation(word_path, excel_path)
    assert validation["errors"] == []

    apply_updates(
        str(excel_path),
        str(output_path),
        updates,
        report_month=word_data["report_month"],
        word_hash="hash",
        excel_hash="hash",
        counts={"metrics": 1, "cases": 1, "articles": 1},
        summary="test",
    )

    wb = load_workbook(output_path)
    ws = wb["Отчет 1-Е Р.2"]
    row = find_row_by_label(ws, "т. 2.1")
    assert row is not None
    assert ws.cell(row=row, column=3).value == 5

    article_row = find_row_by_label(ws, "ст.154")
    assert article_row is not None
    assert ws.cell(row=article_row, column=19).value == 1

