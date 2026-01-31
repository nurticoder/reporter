from __future__ import annotations

import hashlib
from datetime import datetime
from typing import Any

import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def normalize_label(label: str) -> str:
    if label is None:
        return ""
    cleaned = (
        str(label)
        .replace("\u00a0", " ")
        .replace("–", "-")
        .replace("—", "-")
        .strip()
        .lower()
    )
    return "".join(cleaned.split())


def normalize_sheet_name(name: str) -> str:
    if name is None:
        return ""
    cleaned = (
        str(name)
        .replace("\u00a0", " ")
        .replace("–", "-")
        .replace("—", "-")
        .strip()
        .lower()
    )
    cleaned = cleaned.replace("ё", "е")
    cleaned = re.sub(r"\s+", " ", cleaned)
    return re.sub(r"[\s\.-]+", "", cleaned)


def resolve_sheet_name(workbook, desired: str) -> str | None:
    if desired in workbook.sheetnames:
        return desired
    target = normalize_sheet_name(desired)
    partial_matches = []
    for name in workbook.sheetnames:
        normalized = normalize_sheet_name(name)
        if normalized == target:
            return name
        if target and (normalized.endswith(target) or target.endswith(normalized) or target in normalized):
            partial_matches.append(name)
    if len(partial_matches) == 1:
        return partial_matches[0]
    return None


def find_row_by_label(
    ws,
    label: str | None,
    col_letter: str = "B",
    label_contains: str | None = None,
    label_regex: str | None = None,
) -> int | None:
    target = normalize_label(label) if label else ""
    contains_target = normalize_label(label_contains) if label_contains else ""
    regex = re.compile(label_regex, re.IGNORECASE) if label_regex else None
    matches = []
    for row in range(1, ws.max_row + 1):
        value = ws[f"{col_letter}{row}"].value
        normalized_value = normalize_label(value)
        if regex and value is not None and regex.search(str(value)):
            matches.append(row)
        elif contains_target and contains_target in normalized_value:
            matches.append(row)
        elif target and normalized_value == target:
            matches.append(row)
    if len(matches) > 1:
        raise ValueError(f"Multiple rows matched label '{label}' in sheet {ws.title}.")
    return matches[0] if matches else None


def find_column_by_header(ws, header: str, header_row: int | None = None) -> int | None:
    target = normalize_label(header)
    rows_to_scan = [header_row] if header_row else range(1, 6)
    for row in rows_to_scan:
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=row, column=col).value
            if normalize_label(value) == target:
                return col
    return None


def plan_metric_updates(workbook, metrics: dict, excel_map: dict) -> tuple[list[dict], list[dict]]:
    updates = []
    errors = []

    for metric_key, mapping in excel_map.items():
        if metric_key not in metrics:
            continue
        sheet_name = mapping.get("sheet")
        row_label = mapping.get("rowLabel")
        col = mapping.get("col")
        header_row = mapping.get("headerRow")

        resolved_sheet = resolve_sheet_name(workbook, sheet_name)
        if not resolved_sheet:
            errors.append(
                {
                    "type": "error",
                    "message": f"Sheet '{sheet_name}' not found for metric {metric_key}.",
                    "source": "excelMap.json",
                    "suggestedFix": "Update excelMap.json with the correct sheet name.",
                }
            )
            continue

        ws = workbook[resolved_sheet]
        try:
            row = find_row_by_label(
                ws,
                row_label,
                col_letter=mapping.get("rowLabelColumn", "B"),
                label_contains=mapping.get("rowLabelContains"),
                label_regex=mapping.get("rowLabelRegex"),
            )
        except ValueError as exc:
            errors.append(
                {
                    "type": "error",
                    "message": str(exc),
                    "source": sheet_name,
                    "suggestedFix": "Ensure row labels are unique.",
                }
            )
            continue

        if row is None:
            errors.append(
                {
                    "type": "error",
                    "message": f"Row label '{row_label}' not found for metric {metric_key}.",
                    "source": sheet_name,
                    "suggestedFix": "Update excelMap.json row labels to match column B.",
                }
            )
            continue

        if isinstance(col, str):
            col_index = find_column_by_header(ws, col, header_row=header_row)
        else:
            col_index = col
        if not col_index:
            errors.append(
                {
                    "type": "error",
                    "message": f"Column '{col}' not found for metric {metric_key}.",
                    "source": sheet_name,
                    "suggestedFix": "Update excelMap.json with the correct column header or index.",
                }
            )
            continue

        cell = ws.cell(row=row, column=col_index)
        updates.append(
            {
                "sheet": sheet_name,
                "cell": f"{get_column_letter(col_index)}{row}",
                "rowLabel": row_label,
                "oldValue": cell.value,
                "newValue": metrics[metric_key]["value"],
                "kind": metric_key,
                "row": row,
                "col": col_index,
            }
        )

    return updates, errors


def plan_article_updates(workbook, article_breakdown: list[dict], article_map: dict) -> tuple[list[dict], list[dict]]:
    updates = []
    errors = []

    sheet_name = article_map.get("sheet")
    if not sheet_name:
        return updates, errors

    resolved_sheet = resolve_sheet_name(workbook, sheet_name)
    if not resolved_sheet:
        errors.append(
            {
                "type": "error",
                "message": f"Sheet '{sheet_name}' not found for article breakdown.",
                "source": "articleMap.json",
                "suggestedFix": "Update articleMap.json with the correct sheet name.",
            }
        )
        return updates, errors

    ws = workbook[resolved_sheet]
    row_label_column = article_map.get("rowLabelColumn", "B")
    columns = article_map.get("columns", {})

    for row in article_breakdown:
        article = row.get("article")
        if not article:
            continue
        try:
            excel_row = find_row_by_label(
                ws,
                article,
                col_letter=row_label_column,
                label_contains=article_map.get("rowLabelContains"),
                label_regex=article_map.get("rowLabelRegex"),
            )
        except ValueError as exc:
            errors.append(
                {
                    "type": "error",
                    "message": str(exc),
                    "source": sheet_name,
                    "suggestedFix": "Ensure article rows are unique.",
                }
            )
            continue

        if excel_row is None:
            errors.append(
                {
                    "type": "error",
                    "message": f"Article row '{article}' not found in Excel.",
                    "source": sheet_name,
                    "suggestedFix": "Add the article row or update articleMap.json.",
                }
            )
            continue

        for key, col_index in columns.items():
            if key not in row:
                continue
            cell = ws.cell(row=excel_row, column=col_index)
            updates.append(
                {
                    "sheet": sheet_name,
                    "cell": f"{get_column_letter(col_index)}{excel_row}",
                    "rowLabel": article,
                    "oldValue": cell.value,
                    "newValue": row[key],
                    "kind": f"article_{key}",
                    "row": excel_row,
                    "col": col_index,
                }
            )

    return updates, errors


def plan_updates(excel_path: str, metrics: dict, article_breakdown: list[dict], excel_map: dict, article_map: dict):
    workbook = load_workbook(excel_path)
    metric_updates, metric_errors = plan_metric_updates(workbook, metrics, excel_map)
    article_updates, article_errors = plan_article_updates(workbook, article_breakdown, article_map)

    updates = metric_updates + article_updates
    errors = metric_errors + article_errors

    return updates, errors


def append_import_log(
    workbook,
    report_month: dict,
    word_hash: str,
    excel_hash: str,
    status: str,
    counts: dict[str, int],
    summary: str,
):
    sheet_name = "ImportsLog"
    if sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
    else:
        ws = workbook.create_sheet(sheet_name)
        ws.append(
            [
                "timestamp",
                "report_month",
                "word_hash",
                "excel_hash",
                "status",
                "metrics_count",
                "case_count",
                "article_count",
                "summary",
            ]
        )
    ws.append(
        [
            datetime.now().isoformat(timespec="seconds"),
            report_month.get("label") if report_month else None,
            word_hash,
            excel_hash,
            status,
            counts.get("metrics"),
            counts.get("cases"),
            counts.get("articles"),
            summary,
        ]
    )


def sha256_file(path: str) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(8192), b""):
            digest.update(chunk)
    return digest.hexdigest()


def apply_updates(
    excel_path: str,
    output_path: str,
    updates: list[dict],
    report_month: dict,
    word_hash: str,
    excel_hash: str,
    counts: dict[str, int],
    summary: str,
) -> None:
    workbook = load_workbook(excel_path)
    for update in updates:
        ws = workbook[update["sheet"]]
        ws.cell(row=update["row"], column=update["col"], value=update["newValue"])

    append_import_log(
        workbook,
        report_month=report_month,
        word_hash=word_hash,
        excel_hash=excel_hash,
        status="success",
        counts=counts,
        summary=summary,
    )

    workbook.save(output_path)

